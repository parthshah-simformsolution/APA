from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import time
import openpyxl
import sys
import numpy as np
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions




# Create a new instance of the Chrome web driver
driver = webdriver.Chrome()
driver.get('https://genaipoc.apa.org/');
time.sleep(10)
search_question = driver.find_element(By.XPATH, '//div[@data-baseweb="base-input"]/textarea')

search_question.send_keys("memory")
time.sleep(10)
driver.find_element(By.CSS_SELECTOR, 'svg.eyeqlp51.css-9ilocf.ex0cdmw0').click()
time.sleep(10)
j=2
get_ans =[]
xpath_1="(//div[@class='stChatMessage css-4oy321 eeusbqq4']//div[@data-testid='stMarkdownContainer']//p)[" + str(j) + "]"
get_ans=driver.find_element(By.XPATH,xpath_1).text
#print(get_ans)

# Open the spreadsheet
workbook = openpyxl.load_workbook(r"C:\Users\parth\Downloads\APAQuestions.xlsx")

# Get the first sheet
sheet = workbook.worksheets[0]

ans= []
ans.append(get_ans)
print(ans)

search_question.send_keys("gender difference")
time.sleep(10)
driver.find_element(By.CSS_SELECTOR, 'svg.eyeqlp51.css-9ilocf.ex0cdmw0').click()
time.sleep(10)
j=j+1
get_ans=driver.find_element(By.XPATH,xpath_1).text
ans.append(get_ans)
print(ans)

#Path to the Excel file
excel_file_path = r'C:\Users\parth\Downloads\Book1.xlsx'

try:
    for i, answer in enumerate(ans):
        sheet.cell(row=i + 2, column=3, value=answer)

    # Save the modified Excel file
    workbook.save(excel_file_path)
except Exception as e:
    print(f"Error writing answers to Excel: {e}")



# driver.find_element(By.XPATH, '//div[@value="0"]').click()
# driver.find_element(By.XPATH, '//*[text()="Llama-2-7b-Chat"]').click()
# search_question = driver.find_element(By.XPATH, '//div[@data-baseweb="base-input"]/textarea')
# send_question = driver.find_element(By. CSS_SELECTOR, 'svg.eyeqlp51.css-9ilocf.ex0cdmw0')
# for row in range(1, dataframe1.max_row):
# 	for col in dataframe1.iter_cols(2, 2):
# 		a[i]=col[row].value
# 		#data_to_send = a[i]
# 		data_to_send = "".join(a)
# 		print(data_to_send)
# 		time.sleep(20)
# 		search_question.send_keys(data_to_send)
# 		send_question.click()
# 		time.sleep(50)
# 		# j: int = 9
# 		# j = j-1
# 		# ans1 = "(//div[@class='element-container css-onjmxl e1f1d6gn2']//div[@class='stMarkdown']//div[@data-testid='stMarkdownContainer']//p)["
# 		# ans2 = "]"
# 		# ans_xpath_old = ans1 + str(j)+ ans2
# 		# ans_xpath = '"' + ans_xpath_old + '"'
# 		# print(ans_xpath)
# 		# get_ans=driver.find_element(By.XPATH, ans_xpath_old)
# 		# text_content = get_ans.text
# 		# print(text_content)
# try:
#     while True:
#         user_input = input("Press 'q' to quit the browser: ")
#         if user_input.lower() == 'q':
#             break
#
# except KeyboardInterrupt:
#     pass  # Handle keyboard interrupt (e.g., Ctrl+C) gracefully
#
# finally:
#     # Quit the browser
#     driver.quit()



