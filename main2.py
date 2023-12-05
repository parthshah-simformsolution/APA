from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import time
import openpyxl
import sys
import numpy as np
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions




# Create a new instance of the Chrome web driver
driver = webdriver.Chrome()

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(r"C:\Users\parth\Downloads\Book1.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

#Iterate the loop to read the cell values
i: int=0
a=["empty"]

# Navigate to the website
driver.get('https://genaipoc.apa.org/');
driver.implicitly_wait("6")
driver.find_element(By.XPATH, '//div[@value="0"]').click()
driver.find_element(By.XPATH, '//*[text()="Falcon 7B Instruct BF16"]').click()
time.sleep(5)
#driver.find_element(By.XPATH, '//textarea[@name=""]').send_keys("ok go")
search_question = driver.find_element(By.XPATH, '//div[@data-baseweb="base-input"]/textarea')
send_question = driver.find_element(By. CSS_SELECTOR, 'svg.eyeqlp51.css-9ilocf.ex0cdmw0')
#collect_ans = driver.find_element(By.XPATH, '(//div[@class='element-container css-onjmxl e1f1d6gn2']//div[@class='stMarkdown']//div[@data-testid="stMarkdownContainer"]//p)[ + j + ])

for row in range(1, dataframe1.max_row):
	for col in dataframe1.iter_cols(48, 2):
		a[i]=col[row].value
		#data_to_send = a[i]
		data_to_send = "".join(a)
		print(data_to_send)
		time.sleep(3)
		search_question.send_keys(data_to_send)
		send_question.click()
		time.sleep(10)
		# j: int = 9
		# j = j-1
		# ans1 = "(//div[@class='element-container css-onjmxl e1f1d6gn2']//div[@class='stMarkdown']//div[@data-testid='stMarkdownContainer']//p)["
		# ans2 = "]"
		# ans_xpath_old = ans1 + str(j)+ ans2
		# ans_xpath = '"' + ans_xpath_old + '"'
		# print(ans_xpath)
		# get_ans=driver.find_element(By.XPATH, ans_xpath_old)
		# text_content = get_ans.text
		# print(text_content)
try:
    while True:
        user_input = input("Press 'q' to quit the browser: ")
        if user_input.lower() == 'q':
            break

except KeyboardInterrupt:
    pass  # Handle keyboard interrupt (e.g., Ctrl+C) gracefully

finally:
    # Quit the browser
    driver.quit()





#driver n =driver.findElement(By.XPATH, '//textarea[@name=""]')
#s = n.getText()
#print(s)
#print(driver.find_element(By.XPATH, '//textarea[@name=""]').text)


#driver.find_element(By.XPATH, '//*[@title="open"]').click()
#driver.implicitly_wait("60")




